import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, PatternFill
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import textwrap
from Utility import *

class ExcelReport:
    def __init__(self, wells):
        self.wells = wells
        self.data = []

    def collect_data(self, const_list):
        """Collect data from parents and their children."""
        self.data.clear()  # Clear previous data
        for well in self.wells:
            well_name = well.well_name
            print(well.well_name)
            wh=find_dataset_by_name(well,'WELL_HEADER')
            if wh:
                values=[]
                for c in const_list:
                    values.append(find_constant_by_name(wh,c).value if find_constant_by_name(wh, c) else None)
                self.data.append(values)

    def format_excel(self, worksheet):
        """Format the header row and cells with Yes/No values."""
        # Define fill colors
        yes_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for "Yes"
        no_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red for "No"
        
        # Define the border style
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
         # Define a fill for the header background color
        header_fill = PatternFill(start_color="FFFF00",  # Yellow color
                                    end_color="FFFF00",
                                    fill_type="solid")
        # Make the header bold and center-aligned
        for cell in worksheet[1]:  # The first row is the header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border  # Add border to header cells
            cell.fill = header_fill     # Set the background color for header cells

        # Set column widths and add borders around each filled cell
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                    cell.border = thin_border  # Add border to each filled cell
                except Exception as e:
                    pass
            adjusted_width = (max_length + 2)  # Add some padding
            worksheet.column_dimensions[column_letter].width = adjusted_width

            # Optional: You can also add borders to empty cells if needed
            for cell in column:
                if cell.value is None:
                    cell.border = thin_border  # Add border to empty cells

        # Color the cells in the "Has Child with Attribute Value" column
        for row in range(2, len(self.data) + 2):  # Start from the second row to skip the header
            cell = worksheet.cell(row=row, column=4)  # 4th column corresponds to "Has Child with Attribute Value"
            if cell.value == "Yes":
                cell.fill = yes_fill
            elif cell.value == "No":
                cell.fill = no_fill

    def save_to_excel(self, output_file,const_list):
        """Save collected data to an Excel file with formatted cells."""
        # Create a DataFrame
        df = pd.DataFrame(self.data, columns=const_list)
        
        # Check if 'date' is a substring of any column name in a case-insensitive manner
        if any('date' in col.lower() for col in df.columns):
            date_col = [col for col in df.columns if 'date' in col.lower()][0]  # Get the first matching column
            df[date_col] = pd.to_datetime(df[date_col]).dt.strftime('%d-%m-%Y')
            
        print(df)
        # Create a Pandas Excel writer using Openpyxl as the engine
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
            workbook = writer.book
            worksheet = writer.sheets['Data']

            # Call the formatting method
            self.format_excel(worksheet)
            
    def save_to_pdf(self, output_file, const_list):
        """Save collected data to a PDF file."""
        df = pd.DataFrame(self.data, columns=const_list)
        
        # Check if 'date' is a substring of any column name in a case-insensitive manner
        if any('date' in col.lower() for col in df.columns):
            date_col = [col for col in df.columns if 'date' in col.lower()][0]  # Get the first matching column
            df[date_col] = pd.to_datetime(df[date_col]).dt.strftime('%d-%m-%Y')
        
        # Wrap text in each cell
        wrapped_data = [[self.wrap_text(str(cell), width=15) for cell in row] for row in df.values]

        # Calculate column widths based on maximum text length
        column_widths = [max(len(str(cell)) for cell in df[col]) + 2 for col in df.columns]  # Add padding

        # Create a new figure in landscape orientation
        fig, ax = plt.subplots(figsize=(12, 8))  # Width greater than height for landscape


        # Hide axes
        ax.axis('tight')
        ax.axis('off')

        # Create a table and add it to the axes with adjusted column widths
        table = ax.table(cellText=wrapped_data, colLabels=df.columns, cellLoc='center', loc='center')

        # Set column widths based on the max length of the wrapped text
        for i in range(len(df.columns)):
            max_len = max(len(str(cell)) for cell in df.iloc[:, i])
            table.auto_set_column_width(i)
            # Use the max_len to set an appropriate width, adjusting it as needed
            table[(0, i)].set_width((max_len + 2) / 10.0)  # Adjust as needed


        # Style the table
        table.auto_set_font_size(False) 
        table.set_fontsize(10)
        table.scale(1.2, 1.2)

        # Adjust layout
        plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.1)
        
        # Save the figure as a PDF
        with PdfPages(output_file) as pdf:
            pdf.savefig(fig, bbox_inches='tight')
            
    def wrap_text(self, text, width=10):
        """Wrap text to a specified width."""
        if isinstance(text, str):
            return "\n".join(textwrap.wrap(text, width=width))
        return text  # Return as-is if not a string